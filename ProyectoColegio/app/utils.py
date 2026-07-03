"""
UTILIDADES PARA EXPORTACION DE REPORTES
Modulo con funciones para exportar datos a PDF y Excel
"""

from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils import timezone
import io

# ====== EXPORTACION A PDF ======


def exportar_pdf(request, titulo, columnas, datos, nombre_archivo):
    """
    FUNCION PARA EXPORTAR DATOS A PDF USANDO WEASYPRINT

    Args:
        titulo: Titulo del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de tuplas o diccionarios con los datos
        nombre_archivo: Nombre del archivo PDF a descargar

    Returns:
        HttpResponse con el PDF generado
    """
    logo_url = request.build_absolute_uri(static('img/Logo.jpeg'))

    # Crear contexto para el template
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
        'logo_url' : logo_url,
        'now' : timezone.now(),
    }

    # Generar HTML desde el template
    html_string = render_to_string('reportes/reporte_pdf.html', contexto)

    # Crear documento PDF desde el HTML
    html_object = HTML(string=html_string,
                       base_url=request.build_absolute_uri('/'))

    # Generar PDF en memoria
    pdf_bytes = html_object.write_pdf()

    # Crear respuesta HTTP con el PDF
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

    return response


# ====== EXPORTACION A EXCEL ======
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
from django.utils.timezone import is_aware


def exportar_excel(titulo, columnas, datos, nombre_archivo):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"

    # ===== TÍTULO =====
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
    titulo_cell = worksheet['A1']
    titulo_cell.value = titulo
    titulo_cell.font = title_font
    titulo_cell.fill = title_fill
    titulo_cell.alignment = title_alignment
    worksheet.row_dimensions[1].height = 25

    # ===== ENCABEZADOS =====
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_num, columna in enumerate(columnas, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = columna
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    worksheet.row_dimensions[3].height = 20

    # ===== ESTILO DATOS =====
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    data_fill_alternated = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # ===== INSERTAR DATOS =====
    for row_num, fila in enumerate(datos, 4):

        if isinstance(fila, dict):
            valores = [
                fila.get(col.lower().replace(' ', '_'), '')
                for col in columnas
            ]
        else:
            valores = fila

        for col_num, valor in enumerate(valores, 1):
            cell = worksheet.cell(row=row_num, column=col_num)

            # 🔥 CONVERSIÓN SEGURA COMPLETA
            if valor is None:
                valor = ''

            elif isinstance(valor, datetime):
                if is_aware(valor):
                    valor = valor.replace(tzinfo=None)

            elif isinstance(valor, (int, float, str, bool, date)):
                pass

            else:
                valor = str(valor)  # objetos Django

            cell.value = valor
            cell.alignment = data_alignment
            cell.border = data_border

            if (row_num - 4) % 2 == 0:
                cell.fill = data_fill_alternated

    # ===== AJUSTAR COLUMNAS =====
    for col_num, columna in enumerate(columnas, 1):
        max_length = len(str(columna))
        column_letter = chr(64 + col_num)

        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # ===== RESPUESTA HTTP =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'

    workbook.save(response)

    return response
from django.contrib.auth.models import Group

def asignar_rol(usuario, rol):
    grupo = Group.objects.get(name=rol)
    usuario.groups.add(grupo)
    


from django.urls import URLResolver, URLPattern, get_resolver

def obtener_rutas(resolver=None, base=''):
    if resolver is None:
        resolver = get_resolver()
    
    rutas = []

    for pattern in resolver.url_patterns:
        if isinstance(pattern, URLPattern):
            ruta_completa = base + str(pattern.pattern)
            
            if not ruta_completa.startswith('admin/'):
                if not ruta_completa.startswith('/'):
                    ruta_completa = '/' + ruta_completa
                

                ruta_completa = ruta_completa.replace('^', '').replace('$', '')

                if ruta_completa and ruta_completa != '/':
                    rutas.append(ruta_completa)

        elif isinstance(pattern, URLResolver):
            rutas += obtener_rutas(pattern, base + str(pattern.pattern))
    
    return rutas
